from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(r"d:\dev\family-exam-helper")
HEADERS = ["科目", "标签", "题型", "难度", "分值", "题干", "选项A", "选项B", "选项C", "选项D", "正确答案", "解析"]
COLUMN_WIDTHS = {
    "A": 12,
    "B": 16,
    "C": 18,
    "D": 8,
    "E": 10,
    "F": 42,
    "G": 18,
    "H": 18,
    "I": 18,
    "J": 18,
    "K": 12,
    "L": 56,
}

MINIMAL_ROWS = [
    ["数学", "函数", "single_choice", 2, 5, "求 f(x)=x^2 的导数。", "2x", "x^2", "x", "2", "A", "幂函数求导，x^2 的导数为 2x。"],
    ["计算机", "循环", "multi_choice", 2, 6, "以下哪些属于 Python 的循环语句？", "for", "while", "if", "switch", "A,B", "for 和 while 都是循环；if 是条件判断，switch 不是 Python 关键字。"],
    ["通识", "判断", "true_false", 1, "", "地球围绕太阳公转。", "", "", "", "", "A", "判断题会自动生成 A=正确、B=错误；本题答案应为 A。"],
    ["语文", "词语", "fill_blank", 1, "", "请填写成语：锲而不____。", "", "", "", "", "舍", "填空题无需填写选项 A-D，正确答案直接写文本即可。"],
    ["英语", "短语", "fill_blank", 2, 4, "请填写短语：look forward to ____ from you.", "", "", "", "", "hearing", "填空题支持直接填写文本答案。"],
]

EXTENDED_ROWS = [
    ["数学", "函数,导数", "single_choice", 1, 2, "函数 y=x^3 在 x=1 处的导数是多少？", "1", "2", "3", "4", "C", "y'=3x^2，代入 x=1 得 3。"],
    ["数学", "几何,面积", "single_choice", 2, "", "半径为 3 的圆面积是多少？", "6pi", "9pi", "12pi", "18pi", "B", "圆面积公式为 S=pi r^2。"],
    ["数学", "方程,一元二次", "single_choice", 3, 4, "方程 x^2-5x+6=0 的解是？", "1 和 6", "2 和 3", "3 和 5", "1 和 5", "B", "因式分解为 (x-2)(x-3)=0。"],
    ["语文", "成语,积累", "fill_blank", 1, "", "请填写成语：专心____。", "", "", "", "", "致志", "常见成语为专心致志。"],
    ["语文", "古诗,默写", "fill_blank", 2, 3, "补全诗句：海内存知己，天涯若比____。", "", "", "", "", "邻", "完整诗句为天涯若比邻。"],
    ["语文", "病句,辨析", "true_false", 2, "", "“通过这次学习，使我提高了成绩。”这句话没有语病。", "", "", "", "", "B", "该句缺少主语，属于病句。"],
    ["英语", "词汇,时态", "single_choice", 1, 2, "She ____ to school every day.", "go", "goes", "going", "gone", "B", "主语为第三人称单数，动词用 goes。"],
    ["英语", "阅读,固定搭配", "fill_blank", 2, "", "Please ____ the door when you leave.", "", "", "", "", "close", "句意为离开时请关门。"],
    ["英语", "语法,介词", "multi_choice", 2, 5, "以下哪些词可以作介词？", "in", "on", "quickly", "under", "A,B,D", "in、on、under 都可作介词，quickly 是副词。"],
    ["英语", "判断,常识", "true_false", 1, "", "English has 26 letters.", "", "", "", "", "A", "英语字母表共有 26 个字母。"],
    ["计算机", "Python,基础", "single_choice", 1, "", "Python 中用于定义函数的关键字是？", "func", "def", "function", "lambda", "B", "标准函数定义关键字为 def。"],
    ["计算机", "Python,循环", "multi_choice", 2, 4, "以下哪些属于 Python 容器类型？", "list", "dict", "set", "print", "A,B,C", "list、dict、set 都是容器，print 是函数。"],
    ["计算机", "网络,协议", "single_choice", 2, "", "用于网页传输的常见协议是？", "FTP", "SMTP", "HTTP", "SSH", "C", "网页浏览最常见的是 HTTP/HTTPS。"],
    ["计算机", "安全,判断", "true_false", 3, 2, "强密码通常应包含大小写字母、数字和符号。", "", "", "", "", "A", "这是密码安全的基本建议。"],
    ["通识", "科学,常识", "true_false", 1, "", "水在标准大气压下 100 摄氏度沸腾。", "", "", "", "", "A", "这是基础物理常识。"],
    ["通识", "地理,国家", "single_choice", 1, 2, "中国的首都是哪座城市？", "上海", "北京", "广州", "深圳", "B", "中国首都是北京。"],
    ["通识", "历史,事件", "multi_choice", 3, "", "以下哪些属于中国古代四大发明？", "造纸术", "火药", "指南针", "显微镜", "A,B,C", "显微镜不属于四大发明。"],
    ["物理", "力学,单位", "single_choice", 2, 3, "力的国际单位是？", "牛顿", "焦耳", "瓦特", "帕斯卡", "A", "力的单位是牛顿。"],
    ["化学", "元素,符号", "fill_blank", 1, "", "请填写元素符号：氧是 ____。", "", "", "", "", "O", "氧元素符号为 O。"],
    ["生物", "人体,常识", "single_choice", 2, 3, "人体中负责输送血液的器官是？", "肺", "胃", "心脏", "肝脏", "C", "心脏负责泵送血液。"],
]


def create_workbook(output_name: str, title: str, rows: list[list[object]], extra_notes: list[str]) -> Path:
    out = BASE_DIR / output_name
    wb = Workbook()
    ws = wb.active
    ws.title = title

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    note_row = len(rows) + 3
    ws.cell(note_row, 1, "说明")
    ws.cell(note_row, 2, "支持题型：single_choice / multi_choice / true_false / fill_blank")
    for index, note in enumerate(extra_notes, start=1):
        ws.cell(note_row + index, 2, note)

    wb.save(out)
    return out


def main() -> None:
    outputs = [
        create_workbook(
            "excel_import_sample_all_types.xlsx",
            "题库最小示例",
            MINIMAL_ROWS,
            [
                "判断题答案使用 A 或 B，其中 A=正确，B=错误。",
                "填空题无需填写选项 A-D，正确答案直接写文本。",
                "分值列可留空，留空时导入后使用配置文件中的题型默认分值。",
                "Excel 导入适合纯文本题目；公式、图片、代码题建议在网页端编辑。",
            ],
        ),
        create_workbook(
            "excel_import_template_extended.xlsx",
            "题库扩展示例",
            EXTENDED_ROWS,
            [
                "标签列可以写多个标签，当前系统会按原样保存文本，例如：函数,导数。",
                "难度建议使用 1、2、3；导入时其他值不会阻止保存，但建议保持一致。",
                "分值列可手工填写；留空时会使用题型默认分值。",
                "多选题答案用英文逗号分隔，例如 A,B,D。",
                "本模板共 20 道题，可直接导入后再批量修改为你的题库内容。",
            ],
        ),
    ]

    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
