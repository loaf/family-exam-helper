"""
家庭考试助手 - Family Exam Helper
单机运行的考试/练习系统，支持富文本题目、数学公式、代码高亮
"""
import os
import re
import io
import posixpath
import shutil
import sqlite3
import uuid
import json
import time
import zipfile
from urllib.parse import quote, unquote
from datetime import datetime
from functools import wraps

from flask import (Flask, g, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_from_directory,
                   send_file, make_response)
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename
import openpyxl

app = Flask(__name__)
app.secret_key = 'family-exam-helper-2026-sisyphus'
app.config['MAX_FORM_MEMORY_SIZE'] = 8 * 1024 * 1024
app.config['MAX_CONTENT_LENGTH'] = 24 * 1024 * 1024

# ── Directories ──────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BANKS_DIR = os.path.join(BASE_DIR, 'banks')
UPLOADS_DIR = os.path.join(BASE_DIR, 'uploads')
BANK_ASSETS_DIR = os.path.join(BASE_DIR, 'bank_assets')
IMPORTS_DIR = os.path.join(BASE_DIR, 'imports')
SCORING_CONFIG_PATH = os.path.join(BASE_DIR, 'scoring_config.json')

for d in [BANKS_DIR, UPLOADS_DIR, BANK_ASSETS_DIR, IMPORTS_DIR]:
    os.makedirs(d, exist_ok=True)

# ── Allowed extensions ───────────────────────────────────────────────
ALLOWED_IMG_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg', 'bmp'}
MAX_IMG_SIZE = 16 * 1024 * 1024  # 16 MB
DEFAULT_SCORING_CONFIG = {
    'default_scores_by_type': {
        'single_choice': 2,
        'multi_choice': 3,
        'true_false': 1,
        'fill_blank': 2,
    },
    'allow_manual_question_score': True,
    'composite_scoring_mode': 'per_child',
}
_SCORING_CONFIG_CACHE = {
    'mtime': None,
    'data': None,
}


# ══════════════════════════════════════════════════════════════════════
#  Database helpers
# ══════════════════════════════════════════════════════════════════════

def _bank_path(filename):
    """Resolve bank path with traversal protection."""
    path = os.path.realpath(os.path.join(BANKS_DIR, filename))
    if not path.startswith(os.path.realpath(BANKS_DIR)):
        return None
    return path


def _bank_asset_key(bank_filename):
    return os.path.splitext(os.path.basename(bank_filename or ''))[0]


def _bank_assets_dir(bank_filename):
    asset_dir = os.path.realpath(os.path.join(BANK_ASSETS_DIR, _bank_asset_key(bank_filename)))
    if not asset_dir.startswith(os.path.realpath(BANK_ASSETS_DIR)):
        return None
    return asset_dir


def _ensure_bank_assets_dir(bank_filename):
    asset_dir = _bank_assets_dir(bank_filename)
    if asset_dir:
        os.makedirs(asset_dir, exist_ok=True)
    return asset_dir


def _current_bank_filename():
    return session.get('current_bank', '')


def _asset_url(bank_filename, filename):
    return url_for('bank_asset_file', bank_filename=bank_filename, filename=filename)


def _coerce_positive_score(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    if score <= 0:
        return None
    return score


def _load_scoring_config():
    try:
        mtime = os.path.getmtime(SCORING_CONFIG_PATH)
    except OSError:
        mtime = None

    if _SCORING_CONFIG_CACHE['data'] is not None and _SCORING_CONFIG_CACHE['mtime'] == mtime:
        return _SCORING_CONFIG_CACHE['data']

    config = dict(DEFAULT_SCORING_CONFIG)
    config['default_scores_by_type'] = dict(DEFAULT_SCORING_CONFIG['default_scores_by_type'])
    try:
        with open(SCORING_CONFIG_PATH, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        if isinstance(raw, dict):
            defaults = raw.get('default_scores_by_type', {})
            if isinstance(defaults, dict):
                for qtype, fallback in DEFAULT_SCORING_CONFIG['default_scores_by_type'].items():
                    config['default_scores_by_type'][qtype] = _coerce_positive_score(defaults.get(qtype)) or fallback
            config['allow_manual_question_score'] = bool(
                raw.get('allow_manual_question_score', DEFAULT_SCORING_CONFIG['allow_manual_question_score'])
            )
            composite_mode = raw.get('composite_scoring_mode')
            if composite_mode in {'per_child'}:
                config['composite_scoring_mode'] = composite_mode
    except Exception:
        pass

    _SCORING_CONFIG_CACHE['mtime'] = mtime
    _SCORING_CONFIG_CACHE['data'] = config
    return config


def _get_default_score_for_type(qtype):
    config = _load_scoring_config()
    defaults = config.get('default_scores_by_type', {})
    fallback = DEFAULT_SCORING_CONFIG['default_scores_by_type'].get(qtype, 1)
    return float(_coerce_positive_score(defaults.get(qtype)) or fallback)


def _resolve_question_score(question):
    if not question or question.get('type') == 'composite':
        return 0.0, False
    manual_score = _coerce_positive_score(question.get('score'))
    if manual_score is not None:
        return manual_score, False
    return _get_default_score_for_type(question.get('type')), True


def _calculate_question_total_score(question):
    if not question:
        return 0.0
    if question.get('type') != 'composite':
        score, _ = _resolve_question_score(question)
        return score
    return sum(_resolve_question_score(child)[0] for child in question.get('children', []))


def _normalize_asset_match(bank_filename, filename):
    filename = os.path.basename(unquote(filename or ''))
    if not filename:
        return None
    bank_filename = os.path.basename(unquote(bank_filename or ''))
    if bank_filename and bank_filename.endswith('.db'):
        return ('bank', bank_filename, filename)
    return ('legacy', '', filename)


def _extract_asset_refs_from_text(text):
    refs = []
    if not text:
        return refs
    pattern = re.compile(
        r'/(?:uploads/(?P<legacy>[^"\']+)|bank-assets/(?P<bank>[^/]+)/(?P<file>[^"\']+))'
    )
    for match in pattern.finditer(text):
        if match.group('legacy'):
            ref = _normalize_asset_match('', match.group('legacy'))
        else:
            ref = _normalize_asset_match(match.group('bank'), match.group('file'))
        if ref:
            refs.append(ref)
    return refs


def _replace_asset_refs_in_text(text, replacer):
    if not text:
        return text

    def _sub(match):
        if match.group('legacy'):
            replacement = replacer(_normalize_asset_match('', match.group('legacy')))
        else:
            replacement = replacer(_normalize_asset_match(match.group('bank'), match.group('file')))
        return replacement or match.group(0)

    pattern = re.compile(
        r'/(?:uploads/(?P<legacy>[^"\']+)|bank-assets/(?P<bank>[^/]+)/(?P<file>[^"\']+))'
    )
    return pattern.sub(_sub, text)


def _collect_asset_refs_from_payload(question):
    refs = []

    def collect_text(value):
        refs.extend(_extract_asset_refs_from_text(value))

    def collect_options(options):
        for option in options or []:
            collect_text(option.get('content', ''))

    collect_text(question.get('content', ''))
    collect_text(question.get('explanation', ''))
    collect_options(question.get('options', []))
    for child in question.get('children', []) or []:
        collect_text(child.get('content', ''))
        collect_text(child.get('explanation', ''))
        collect_options(child.get('options', []))

    unique = []
    seen = set()
    for ref in refs:
        if ref not in seen:
            seen.add(ref)
            unique.append(ref)
    return unique


def _resolve_asset_source(ref):
    ref_type, bank_filename, filename = ref
    if ref_type == 'bank':
        asset_dir = _bank_assets_dir(bank_filename)
        if not asset_dir:
            return None
        return os.path.join(asset_dir, filename)
    return os.path.join(UPLOADS_DIR, filename)


def _copy_asset_to_bank(ref, target_bank_filename):
    source_path = _resolve_asset_source(ref)
    if not source_path or not os.path.exists(source_path):
        return None

    asset_dir = _ensure_bank_assets_dir(target_bank_filename)
    ext = os.path.splitext(ref[2])[1].lower()
    new_filename = f"{uuid.uuid4().hex}{ext}"
    target_path = os.path.join(asset_dir, new_filename)
    shutil.copy2(source_path, target_path)
    return new_filename


def _remap_question_assets(question, target_bank_filename, copied_assets=None):
    copied_assets = copied_assets if copied_assets is not None else {}

    def replacer(ref):
        if not ref:
            return None
        if ref not in copied_assets:
            copied_assets[ref] = _copy_asset_to_bank(ref, target_bank_filename)
        new_filename = copied_assets.get(ref)
        if not new_filename:
            return None
        return _asset_url(target_bank_filename, new_filename)

    def remap_text(value):
        return _replace_asset_refs_in_text(value, replacer)

    def remap_options(options):
        mapped = []
        for option in options or []:
            item = dict(option)
            item['content'] = remap_text(item.get('content', ''))
            mapped.append(item)
        return mapped

    mapped = dict(question)
    mapped['content'] = remap_text(mapped.get('content', ''))
    mapped['explanation'] = remap_text(mapped.get('explanation', ''))
    mapped['options'] = remap_options(mapped.get('options', []))
    if mapped.get('type') == 'composite':
        mapped['children'] = []
        for child in question.get('children', []) or []:
            child_item = dict(child)
            child_item['content'] = remap_text(child_item.get('content', ''))
            child_item['explanation'] = remap_text(child_item.get('explanation', ''))
            child_item['options'] = remap_options(child_item.get('options', []))
            mapped['children'].append(child_item)
    return mapped


QUESTION_TYPE_LABELS = {
    'single_choice': '单选题',
    'multi_choice': '多选题',
    'true_false': '判断题',
    'fill_blank': '填空题',
    'composite': '组合题',
}


ANSWERABLE_TYPES = ('single_choice', 'multi_choice', 'true_false', 'fill_blank')


def _ensure_question_schema(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id INTEGER DEFAULT NULL,
            tag TEXT DEFAULT '',
            type TEXT NOT NULL,
            difficulty INTEGER DEFAULT 1 CHECK(difficulty BETWEEN 1 AND 3),
            content TEXT NOT NULL DEFAULT '',
            options TEXT DEFAULT '[]',
            answer TEXT NOT NULL DEFAULT '',
            explanation TEXT DEFAULT '',
            score REAL DEFAULT NULL,
            parent_id INTEGER DEFAULT NULL,
            sort_order INTEGER DEFAULT 0,
            stem TEXT DEFAULT '',
            is_subquestion INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE SET NULL,
            FOREIGN KEY (parent_id) REFERENCES questions(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_questions_subject ON questions(subject_id);
        CREATE INDEX IF NOT EXISTS idx_questions_type ON questions(type);
        CREATE INDEX IF NOT EXISTS idx_questions_tag ON questions(tag);
        CREATE INDEX IF NOT EXISTS idx_questions_parent_id ON questions(parent_id);
        CREATE INDEX IF NOT EXISTS idx_questions_parent_sort ON questions(parent_id, sort_order);
    """)


def migrate_bank_schema(conn):
    """Ensure existing banks support composite questions."""
    columns = {
        row[1] for row in conn.execute("PRAGMA table_info(questions)").fetchall()
    }
    needs_rebuild = not {'parent_id', 'sort_order', 'stem', 'is_subquestion'}.issubset(columns)
    if not needs_rebuild:
        _ensure_question_schema(conn)
        if 'score' not in columns:
            conn.execute("ALTER TABLE questions ADD COLUMN score REAL DEFAULT NULL")
        _repair_orphan_composites(conn)
        conn.commit()
        return

    score_select = 'score' if 'score' in columns else 'NULL'
    conn.execute("PRAGMA foreign_keys = OFF")
    conn.executescript("""
        ALTER TABLE questions RENAME TO questions_legacy;

        CREATE TABLE questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id INTEGER DEFAULT NULL,
            tag TEXT DEFAULT '',
            type TEXT NOT NULL,
            difficulty INTEGER DEFAULT 1 CHECK(difficulty BETWEEN 1 AND 3),
            content TEXT NOT NULL DEFAULT '',
            options TEXT DEFAULT '[]',
            answer TEXT NOT NULL DEFAULT '',
            explanation TEXT DEFAULT '',
            score REAL DEFAULT NULL,
            parent_id INTEGER DEFAULT NULL,
            sort_order INTEGER DEFAULT 0,
            stem TEXT DEFAULT '',
            is_subquestion INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE SET NULL,
            FOREIGN KEY (parent_id) REFERENCES questions(id) ON DELETE CASCADE
        );
    """)
    conn.execute(
        """
        INSERT INTO questions
            (id, subject_id, tag, type, difficulty, content, options, answer, explanation, score,
             parent_id, sort_order, stem, is_subquestion, created_at)
        SELECT id, subject_id, tag, type, difficulty, content, options, answer, explanation, """ + score_select + """,
               NULL, 0, content, 0, created_at
        FROM questions_legacy
        """
    )
    conn.executescript("""
        DROP TABLE questions_legacy;
        CREATE INDEX IF NOT EXISTS idx_questions_subject ON questions(subject_id);
        CREATE INDEX IF NOT EXISTS idx_questions_type ON questions(type);
        CREATE INDEX IF NOT EXISTS idx_questions_tag ON questions(tag);
        CREATE INDEX IF NOT EXISTS idx_questions_parent_id ON questions(parent_id);
        CREATE INDEX IF NOT EXISTS idx_questions_parent_sort ON questions(parent_id, sort_order);
    """)
    conn.execute("PRAGMA foreign_keys = ON")
    _repair_orphan_composites(conn)
    conn.commit()


def _repair_orphan_composites(conn):
    """
    Repair composite parents that lost their children due to earlier bad migrations.
    Heuristic: reattach top-level questions created in the same batch timestamp,
    with same subject and tag, and created right after the composite parent.
    """
    parents = conn.execute(
        """
        SELECT id, subject_id, tag, created_at, content
        FROM questions
        WHERE type='composite' AND is_subquestion=0
        ORDER BY id
        """
    ).fetchall()
    repaired = False
    for parent in parents:
        parent_id = parent['id'] if isinstance(parent, sqlite3.Row) else parent[0]
        parent_subject_id = parent['subject_id'] if isinstance(parent, sqlite3.Row) else parent[1]
        parent_tag = parent['tag'] if isinstance(parent, sqlite3.Row) else parent[2]
        parent_created_at = parent['created_at'] if isinstance(parent, sqlite3.Row) else parent[3]
        parent_content = parent['content'] if isinstance(parent, sqlite3.Row) else parent[4]
        child_count = conn.execute(
            "SELECT COUNT(*) FROM questions WHERE parent_id=?",
            (parent_id,)
        ).fetchone()[0]
        if child_count:
            continue

        candidates = conn.execute(
            """
            SELECT id
            FROM questions
            WHERE id > ?
              AND type != 'composite'
              AND parent_id IS NULL
              AND is_subquestion = 0
              AND created_at = ?
              AND COALESCE(subject_id, -1) = COALESCE(?, -1)
              AND tag = ?
            ORDER BY id
            """,
            (parent_id, parent_created_at, parent_subject_id, parent_tag)
        ).fetchall()
        if not candidates:
            continue

        for index, row in enumerate(candidates, start=1):
            child_id = row['id'] if isinstance(row, sqlite3.Row) else row[0]
            conn.execute(
                """
                UPDATE questions
                SET parent_id=?, is_subquestion=1, sort_order=?, stem=?
                WHERE id=?
                """,
                (parent_id, index, parent_content, child_id)
            )
        repaired = True

    if repaired:
        conn.commit()


def _decode_options(raw):
    if not raw:
        return []
    try:
        return json.loads(raw)
    except Exception:
        return []


def _normalize_multi_answer(answer):
    parts = [part.strip().upper() for part in re.split(r'[,，\s]+', answer or '') if part.strip()]
    return ','.join(parts)


def _question_row_to_dict(row):
    data = dict(row)
    data['options'] = _decode_options(row['options'])
    data['score'] = _coerce_positive_score(data.get('score'))
    data['type_label'] = QUESTION_TYPE_LABELS.get(row['type'], row['type'])
    data['is_composite'] = row['type'] == 'composite'
    return data


def _get_bank_display_name(conn, fallback_name='bank'):
    try:
        row = conn.execute("SELECT value FROM bank_meta WHERE key='display_name'").fetchone()
        if row and row[0]:
            return row[0]
    except Exception:
        pass
    return fallback_name


def _build_export_filters(source):
    return {
        'subject': (source.get('subject') or '').strip(),
        'tag': (source.get('tag') or '').strip(),
        'type': (source.get('type') or '').strip(),
        'q': (source.get('q') or '').strip(),
    }


def _query_top_level_questions(db, filters=None):
    filters = filters or {}
    query = (
        "SELECT q.*, s.name AS subject_name "
        "FROM questions q LEFT JOIN subjects s ON q.subject_id=s.id "
        "WHERE q.is_subquestion=0"
    )
    params = []
    if filters.get('subject'):
        query += " AND q.subject_id=?"
        params.append(filters['subject'])
    if filters.get('tag'):
        query += " AND q.tag=?"
        params.append(filters['tag'])
    if filters.get('type'):
        query += " AND q.type=?"
        params.append(filters['type'])
    if filters.get('q'):
        keyword = f"%{filters['q']}%"
        query += (
            " AND ("
            "q.content LIKE ? OR q.explanation LIKE ? OR q.tag LIKE ? "
            "OR EXISTS ("
            "SELECT 1 FROM questions c "
            "WHERE c.parent_id = q.id AND (c.content LIKE ? OR c.explanation LIKE ?)"
            ")"
            ")"
        )
        params.extend([keyword, keyword, keyword, keyword, keyword])
    query += " ORDER BY q.id"
    return db.execute(query, params).fetchall()


def _serialize_question_export(question):
    payload = {
        'subject': question.get('subject_name') or '',
        'tag': question.get('tag', ''),
        'type': question.get('type', 'single_choice'),
        'difficulty': question.get('difficulty', 1),
        'score': question.get('score'),
        'content': question.get('content', ''),
        'options': question.get('options', []),
        'answer': question.get('answer', ''),
        'explanation': question.get('explanation', ''),
    }
    if question.get('type') == 'composite':
        payload['children'] = [
            {
                'type': child.get('type', 'single_choice'),
                'difficulty': child.get('difficulty', 1),
                'score': child.get('score'),
                'content': child.get('content', ''),
                'options': child.get('options', []),
                'answer': child.get('answer', ''),
                'explanation': child.get('explanation', ''),
            }
            for child in question.get('children', [])
        ]
    return payload


def _ensure_subject(existing_subjects, db, subject_name):
    name = (subject_name or '').strip()
    if not name:
        return None
    if name not in existing_subjects:
        cursor = db.execute('INSERT INTO subjects (name) VALUES (?)', (name,))
        existing_subjects[name] = cursor.lastrowid
    return existing_subjects[name]


def _import_question_payload(db, existing_subjects, question):
    subject_id = _ensure_subject(existing_subjects, db, question.get('subject', ''))
    tag = question.get('tag', '')
    qtype = question.get('type', 'single_choice')
    difficulty = int(question.get('difficulty', 1) or 1)
    score = _coerce_positive_score(question.get('score'))
    content = question.get('content', '')
    explanation = question.get('explanation', '')

    if qtype == 'composite':
        children = question.get('children', []) or []
        cursor = db.execute(
            """
            INSERT INTO questions
                (subject_id,tag,type,difficulty,content,options,answer,explanation,score,stem,is_subquestion)
            VALUES (?,?,?,?,?,?,?,?,?,?,0)
            """,
            (subject_id, tag, 'composite', difficulty, content, '[]', '', explanation, None, content)
        )
        parent_id = cursor.lastrowid
        for index, child in enumerate(children, start=1):
            db.execute(
                """
                INSERT INTO questions
                    (subject_id,tag,type,difficulty,content,options,answer,explanation,score,
                     parent_id,sort_order,stem,is_subquestion)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1)
                """,
                (
                    subject_id,
                    tag,
                    child.get('type', 'single_choice'),
                    int(child.get('difficulty', difficulty) or difficulty),
                    child.get('content', ''),
                    json.dumps(child.get('options', []) or [], ensure_ascii=False),
                    child.get('answer', ''),
                    child.get('explanation', ''),
                    _coerce_positive_score(child.get('score')),
                    parent_id,
                    index,
                    content,
                )
            )
        return

    db.execute(
        """
        INSERT INTO questions
            (subject_id,tag,type,difficulty,content,options,answer,explanation,score,stem,is_subquestion)
        VALUES (?,?,?,?,?,?,?,?,?,?,0)
        """,
        (
            subject_id,
            tag,
            qtype,
            difficulty,
            content,
            json.dumps(question.get('options', []) or [], ensure_ascii=False),
            question.get('answer', ''),
            explanation,
            score,
            content,
        )
    )


def _build_export_file_name(bank_name, filters):
    safe_name = bank_name if bank_name else 'bank'
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', safe_name)
    parts = []
    if filters.get('subject_name'):
        parts.append(filters['subject_name'])
    if filters.get('tag'):
        parts.append(filters['tag'])
    if filters.get('type_label'):
        parts.append(filters['type_label'])
    if parts:
        safe_name += '_' + '_'.join(parts)
    return safe_name


def _make_json_download_response(data, file_name):
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', file_name or 'bank')
    ascii_name = safe_name.encode('ascii', 'ignore').decode('ascii') or 'bank'
    resp = make_response(json.dumps(data, ensure_ascii=False, indent=2))
    resp.headers['Content-Type'] = 'application/json; charset=utf-8'
    resp.headers['Content-Disposition'] = (
        f'attachment; filename="{ascii_name}.json"; '
        f"filename*=UTF-8''{quote(safe_name + '.json')}"
    )
    return resp


def _make_zip_download_response(data_bytes, file_name):
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', file_name or 'bank')
    ascii_name = safe_name.encode('ascii', 'ignore').decode('ascii') or 'bank'
    resp = make_response(data_bytes)
    resp.headers['Content-Type'] = 'application/zip'
    resp.headers['Content-Disposition'] = (
        f'attachment; filename="{ascii_name}.zip"; '
        f"filename*=UTF-8''{quote(safe_name + '.zip')}"
    )
    return resp


def _asset_ref_to_manifest_item(ref, archive_name):
    return {
        'type': ref[0],
        'bank_filename': ref[1],
        'filename': ref[2],
        'archive_name': archive_name,
    }


def _manifest_item_to_asset_ref(item):
    return (
        item.get('type', 'legacy'),
        item.get('bank_filename', ''),
        os.path.basename(item.get('filename', '')),
    )


def _build_export_package(data, file_name):
    buffer = io.BytesIO()
    asset_manifest = []
    asset_paths = {}

    for question in data.get('questions', []):
        for ref in _collect_asset_refs_from_payload(question):
            if ref in asset_paths:
                continue
            source_path = _resolve_asset_source(ref)
            if not source_path or not os.path.exists(source_path):
                continue
            archive_name = f"assets/{uuid.uuid4().hex}{os.path.splitext(ref[2])[1].lower()}"
            asset_paths[ref] = archive_name
            asset_manifest.append(_asset_ref_to_manifest_item(ref, archive_name))

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('bank.json', json.dumps(data, ensure_ascii=False, indent=2))
        zf.writestr('manifest.json', json.dumps({
            'format': 'family-exam-helper-package',
            'version': 1,
            'assets': asset_manifest,
        }, ensure_ascii=False, indent=2))
        for ref, archive_name in asset_paths.items():
            zf.write(_resolve_asset_source(ref), archive_name)

    return _make_zip_download_response(buffer.getvalue(), file_name)


def _load_import_payload(uploaded_file):
    filename = (uploaded_file.filename or '').lower()
    if filename.endswith('.json'):
        raw = uploaded_file.read().decode('utf-8')
        return json.loads(raw), {}
    if not filename.endswith('.zip'):
        raise ValueError('请上传导出包（.zip）或 JSON 文件（.json）')

    raw = uploaded_file.read()
    with zipfile.ZipFile(io.BytesIO(raw), 'r') as zf:
        try:
            data = json.loads(zf.read('bank.json').decode('utf-8'))
        except KeyError as exc:
            raise ValueError('导出包缺少 bank.json') from exc

        manifest = {}
        if 'manifest.json' in zf.namelist():
            manifest_data = json.loads(zf.read('manifest.json').decode('utf-8'))
            for item in manifest_data.get('assets', []):
                ref = _manifest_item_to_asset_ref(item)
                archive_name = item.get('archive_name', '')
                if not archive_name or archive_name not in zf.namelist():
                    continue
                ext = os.path.splitext(ref[2])[1].lower()
                new_filename = f"{uuid.uuid4().hex}{ext}"
                asset_dir = _ensure_bank_assets_dir(_current_bank_filename())
                with zf.open(archive_name) as source, open(os.path.join(asset_dir, new_filename), 'wb') as target:
                    shutil.copyfileobj(source, target)
                manifest[ref] = new_filename

    return data, manifest


def _inline_imported_asset_refs(question, imported_assets, target_bank_filename):
    if not imported_assets:
        return question

    def replacer(ref):
        new_filename = imported_assets.get(ref)
        if not new_filename:
            return None
        return _asset_url(target_bank_filename, new_filename)

    def remap_text(value):
        return _replace_asset_refs_in_text(value, replacer)

    def remap_options(options):
        result = []
        for option in options or []:
            item = dict(option)
            item['content'] = remap_text(item.get('content', ''))
            result.append(item)
        return result

    mapped = dict(question)
    mapped['content'] = remap_text(mapped.get('content', ''))
    mapped['explanation'] = remap_text(mapped.get('explanation', ''))
    mapped['options'] = remap_options(mapped.get('options', []))
    if mapped.get('type') == 'composite':
        mapped['children'] = []
        for child in question.get('children', []) or []:
            child_item = dict(child)
            child_item['content'] = remap_text(child_item.get('content', ''))
            child_item['explanation'] = remap_text(child_item.get('explanation', ''))
            child_item['options'] = remap_options(child_item.get('options', []))
            mapped['children'].append(child_item)
    return mapped


def _normalize_template_asset_path(path):
    raw = unquote((path or '').strip()).replace('\\', '/')
    if not raw:
        return None
    normalized = posixpath.normpath(raw)
    if normalized in {'.', ''}:
        return None
    if normalized.startswith('./'):
        normalized = normalized[2:]
    if normalized.startswith('../') or normalized == '..':
        return None
    if normalized.startswith('/'):
        normalized = normalized.lstrip('/')
    if normalized == 'assets' or not normalized.startswith('assets/'):
        return None
    return normalized


def _replace_template_asset_paths_in_text(text, imported_assets, target_bank_filename):
    if not text or not imported_assets:
        return text
    rendered = text
    for asset_path, new_filename in sorted(imported_assets.items(), key=lambda item: len(item[0]), reverse=True):
        target_url = _asset_url(target_bank_filename, new_filename)
        encoded = quote(asset_path, safe='/')
        for candidate in (f'./{asset_path}', asset_path, f'./{encoded}', encoded):
            rendered = rendered.replace(candidate, target_url)
    return rendered


def _inline_template_imported_asset_refs(question, imported_assets, target_bank_filename):
    if not imported_assets:
        return question

    def remap_text(value):
        return _replace_template_asset_paths_in_text(value, imported_assets, target_bank_filename)

    def remap_options(options):
        result = []
        for option in options or []:
            item = dict(option)
            item['content'] = remap_text(item.get('content', ''))
            result.append(item)
        return result

    mapped = dict(question)
    mapped['content'] = remap_text(mapped.get('content', ''))
    mapped['explanation'] = remap_text(mapped.get('explanation', ''))
    mapped['options'] = remap_options(mapped.get('options', []))
    if mapped.get('type') == 'composite':
        mapped['children'] = []
        for child in question.get('children', []) or []:
            child_item = dict(child)
            child_item['content'] = remap_text(child_item.get('content', ''))
            child_item['explanation'] = remap_text(child_item.get('explanation', ''))
            child_item['options'] = remap_options(child_item.get('options', []))
            mapped['children'].append(child_item)
    return mapped


def _extract_template_subject_names(data):
    names = []
    seen = set()
    for item in data.get('subjects', []) or []:
        if isinstance(item, dict):
            name = (item.get('name') or '').strip()
        else:
            name = str(item or '').strip()
        if name and name not in seen:
            seen.add(name)
            names.append(name)

    for question in data.get('questions', []) or []:
        name = (question.get('subject') or '').strip()
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return names


def _load_standard_template_payload(uploaded_file):
    filename = (uploaded_file.filename or '').lower()
    if filename.endswith('.json'):
        raw = uploaded_file.read().decode('utf-8')
        data = json.loads(raw)
        return data, {}
    if not filename.endswith('.zip'):
        raise ValueError('请上传标准模板文件（.zip）或模板 JSON（.json）')

    raw = uploaded_file.read()
    with zipfile.ZipFile(io.BytesIO(raw), 'r') as zf:
        try:
            data = json.loads(zf.read('template.json').decode('utf-8'))
        except KeyError as exc:
            raise ValueError('标准模板压缩包缺少 template.json') from exc

        imported_assets = {}
        asset_dir = _ensure_bank_assets_dir(_current_bank_filename())
        for name in zf.namelist():
            if name.endswith('/'):
                continue
            asset_path = _normalize_template_asset_path(name)
            if not asset_path:
                continue
            ext = os.path.splitext(asset_path)[1].lower()
            new_filename = f"{uuid.uuid4().hex}{ext}"
            with zf.open(name) as source, open(os.path.join(asset_dir, new_filename), 'wb') as target:
                shutil.copyfileobj(source, target)
            imported_assets[asset_path] = new_filename

    return data, imported_assets


def _fetch_children(db, parent_id):
    rows = db.execute(
        """
        SELECT q.*, s.name AS subject_name
        FROM questions q
        LEFT JOIN subjects s ON q.subject_id = s.id
        WHERE q.parent_id=?
        ORDER BY q.sort_order, q.id
        """,
        (parent_id,)
    ).fetchall()
    return [_question_row_to_dict(row) for row in rows]


def _load_question_unit(db, qid):
    row = db.execute(
        """
        SELECT q.*, s.name AS subject_name
        FROM questions q
        LEFT JOIN subjects s ON q.subject_id = s.id
        WHERE q.id=?
        """,
        (qid,)
    ).fetchone()
    if not row:
        return None
    question = _question_row_to_dict(row)
    if question['type'] == 'composite':
        question['children'] = _fetch_children(db, qid)
        question['children_count'] = len(question['children'])
    else:
        question['children'] = []
        question['children_count'] = 0
    return question


def _serialize_answer_for_session(value):
    if isinstance(value, dict):
        return {str(k): v for k, v in value.items()}
    return value or ''


def _parse_question_answer(qtype, raw_answer):
    answer = (raw_answer or '').strip()
    if qtype == 'fill_blank':
        return '|'.join(a.strip() for a in answer.split('|') if a.strip())
    if qtype == 'multi_choice':
        return _normalize_multi_answer(answer)
    return answer


def _build_question_options(qtype, option_map):
    if qtype == 'true_false':
        return [
            {'label': 'A', 'content': '正确'},
            {'label': 'B', 'content': '错误'},
        ]
    if qtype not in ('single_choice', 'multi_choice'):
        return []

    options = []
    for label in ['A', 'B', 'C', 'D', 'E', 'F']:
        val = (option_map.get(label.lower()) or option_map.get(label) or '').strip()
        if val:
            options.append({'label': label, 'content': val})
    return options


def _extract_composite_answers(form, children):
    answers = {}
    for child in children:
        key = f"sub_answer_{child['id']}"
        value = (form.get(key) or '').strip()
        if child['type'] == 'multi_choice':
            value = _normalize_multi_answer(value)
        answers[str(child['id'])] = value
    return answers


def _count_answerable_questions(db):
    row = db.execute(
        "SELECT COUNT(*) FROM questions WHERE type != 'composite'"
    ).fetchone()
    return row[0] if row else 0


def _count_bank_switch_questions(conn):
    columns = {
        row[1] for row in conn.execute("PRAGMA table_info(questions)").fetchall()
    }
    if 'is_subquestion' in columns:
        row = conn.execute(
            "SELECT COUNT(*) FROM questions WHERE COALESCE(is_subquestion, 0)=0"
        ).fetchone()
    else:
        row = conn.execute("SELECT COUNT(*) FROM questions").fetchone()
    return row[0] if row else 0


def _build_result_entry(question, user_answer):
    if question['type'] != 'composite':
        is_correct = _check_answer(question['type'], question['answer'], user_answer)
        max_score, uses_default_score = _resolve_question_score(question)
        return {
            'id': question['id'],
            'type': question['type'],
            'type_label': question['type_label'],
            'content': question['content'],
            'options': question['options'],
            'answer': question['answer'],
            'explanation': question['explanation'],
            'user_answer': user_answer,
            'is_correct': is_correct,
            'difficulty': question['difficulty'],
            'score': max_score if is_correct else 0,
            'max_score': max_score,
            'manual_score': question.get('score'),
            'uses_default_score': uses_default_score,
            'group_correct': 1 if is_correct else 0,
            'group_total': 1,
        }

    children_results = []
    group_correct = 0
    total_score = 0
    max_score = 0
    answers = user_answer if isinstance(user_answer, dict) else {}
    for idx, child in enumerate(question['children'], start=1):
        child_answer = answers.get(str(child['id']), '')
        is_correct = _check_answer(child['type'], child['answer'], child_answer)
        child_max_score, child_uses_default = _resolve_question_score(child)
        max_score += child_max_score
        if is_correct:
            group_correct += 1
            total_score += child_max_score
        children_results.append({
            'id': child['id'],
            'index': idx,
            'type': child['type'],
            'type_label': child['type_label'],
            'content': child['content'],
            'options': child['options'],
            'answer': child['answer'],
            'explanation': child['explanation'],
            'user_answer': child_answer,
            'is_correct': is_correct,
            'difficulty': child['difficulty'],
            'score': child_max_score if is_correct else 0,
            'max_score': child_max_score,
            'manual_score': child.get('score'),
            'uses_default_score': child_uses_default,
        })

    return {
        'id': question['id'],
        'type': question['type'],
        'type_label': question['type_label'],
        'content': question['content'],
        'stem': question.get('stem') or question['content'],
        'answer': '',
        'explanation': question['explanation'],
        'user_answer': '',
        'is_correct': group_correct == len(question['children']) and bool(question['children']),
        'difficulty': question['difficulty'],
        'options': [],
        'children_results': children_results,
        'group_correct': group_correct,
        'group_total': len(question['children']),
        'score': total_score,
        'max_score': max_score,
    }


def init_bank_db(db_path, display_name=''):
    """Create tables for a new question bank."""
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS bank_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS exam_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mode TEXT NOT NULL,
            subject_id INTEGER DEFAULT NULL,
            tag TEXT DEFAULT '',
            total_questions INTEGER DEFAULT 0,
            correct_count INTEGER DEFAULT 0,
            score REAL DEFAULT 0,
            total_score REAL DEFAULT 0,
            duration_seconds INTEGER DEFAULT 0,
            answers TEXT DEFAULT '[]',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

    """)
    _ensure_question_schema(conn)
    if display_name:
        conn.execute("INSERT INTO bank_meta (key, value) VALUES ('display_name', ?)", (display_name,))
    conn.commit()
    conn.close()


def get_bank_list():
    """List all .db banks with display name and question count."""
    banks = []
    for f in sorted(os.listdir(BANKS_DIR)):
        if not f.endswith('.db'):
            continue
        db_path = os.path.join(BANKS_DIR, f)
        name = f[:-3]  # fallback
        count = 0
        try:
            conn = sqlite3.connect(db_path)
            row = conn.execute("SELECT value FROM bank_meta WHERE key='display_name'").fetchone()
            if row and row[0]:
                name = row[0]
            migrate_bank_schema(conn)
            count = _count_bank_switch_questions(conn)
            conn.close()
        except Exception:
            pass
        banks.append({'filename': f, 'name': name, 'count': count})
    return banks


def get_db():
    """Return a per-request sqlite3 connection for the current bank."""
    if 'db' not in g:
        db_name = session.get('current_bank')
        if not db_name:
            return None
        db_path = _bank_path(db_name)
        if not db_path or not os.path.exists(db_path):
            return None
        conn = sqlite3.connect(db_path, detect_types=sqlite3.PARSE_DECLTYPES)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA journal_mode = WAL")
        migrate_bank_schema(conn)
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def require_bank(f):
    """Decorator: redirect to home if no bank selected."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('current_bank') or get_db() is None:
            flash('请先选择题库', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return wrapper


@app.errorhandler(RequestEntityTooLarge)
def handle_request_too_large(exc):
    flash('提交内容过大，请减少题干/解析中的内容体积后重试。', 'error')
    target = request.referrer or url_for('list_questions')
    return redirect(target)


# ══════════════════════════════════════════════════════════════════════
#  Image upload (TinyMCE)
# ══════════════════════════════════════════════════════════════════════

def _allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMG_EXT


@app.route('/upload_image', methods=['POST'])
@require_bank
def upload_image():
    """TinyMCE image upload endpoint. Returns {location: url}."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    if f.filename == '' or not _allowed_image(f.filename):
        return jsonify({'error': 'Invalid file'}), 400

    ext = f.filename.rsplit('.', 1)[1].lower()
    fname = f"{uuid.uuid4().hex}.{ext}"
    bank_filename = _current_bank_filename()
    asset_dir = _ensure_bank_assets_dir(bank_filename)
    f.save(os.path.join(asset_dir, fname))
    return jsonify({'location': _asset_url(bank_filename, fname)})


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOADS_DIR, filename)


@app.route('/bank-assets/<bank_filename>/<filename>')
def bank_asset_file(bank_filename, filename):
    bank_filename = os.path.basename(bank_filename or '')
    if not bank_filename.endswith('.db'):
        bank_filename = f'{bank_filename}.db'
    asset_dir = _bank_assets_dir(bank_filename)
    if not asset_dir or not os.path.exists(os.path.join(asset_dir, filename)):
        return send_from_directory(UPLOADS_DIR, filename)
    return send_from_directory(asset_dir, filename)


# ══════════════════════════════════════════════════════════════════════
#  Routes – Bank management
# ══════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    banks = get_bank_list()
    current = session.get('current_bank')
    return render_template('index.html', banks=banks, current_bank=current)


@app.route('/bank/create', methods=['POST'])
def create_bank():
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入题库名称', 'error')
        return redirect(url_for('index'))
    # Use UUID for filename (avoids encoding issues with Chinese)
    filename = f"{uuid.uuid4().hex}.db"
    while os.path.exists(os.path.join(BANKS_DIR, filename)):
        filename = f"{uuid.uuid4().hex}.db"
    init_bank_db(os.path.join(BANKS_DIR, filename), display_name=name)
    session['current_bank'] = filename
    flash(f'题库「{name}」创建成功', 'success')
    return redirect(url_for('bank_home'))


@app.route('/bank/select/<filename>')
def select_bank(filename):
    if not filename.endswith('.db'):
        flash('无效的题库', 'error')
        return redirect(url_for('index'))
    path = _bank_path(filename)
    if not path or not os.path.exists(path):
        flash('题库不存在', 'error')
        return redirect(url_for('index'))
    session['current_bank'] = filename
    return redirect(url_for('bank_home'))


@app.route('/bank/delete/<filename>', methods=['POST'])
def delete_bank(filename):
    path = _bank_path(filename)
    if not path or not os.path.exists(path):
        flash('题库不存在', 'error')
        return redirect(url_for('index'))
    asset_dir = _bank_assets_dir(filename)

    # Close current connection if it's to this bank
    if session.get('current_bank') == filename:
        db = g.pop('db', None)
        if db is not None:
            db.close()
        session.pop('current_bank', None)

    # Checkpoint WAL and close any lingering connections
    try:
        conn = sqlite3.connect(path)
        conn.execute('PRAGMA wal_checkpoint(TRUNCATE)')
        conn.close()
    except Exception:
        pass

    # Remove the .db file and any WAL/SHM sidecar files
    base = path[:-3] if path.endswith('.db') else path
    for suffix in ['', '-shm', '-wal']:
        f = base + suffix
        if os.path.exists(f):
            try:
                os.remove(f)
            except Exception:
                pass
    if asset_dir and os.path.exists(asset_dir):
        shutil.rmtree(asset_dir, ignore_errors=True)

    flash('题库已删除', 'success')
    return redirect(url_for('index'))


@app.route('/bank/back')
def back_to_banks():
    session.pop('current_bank', None)
    return redirect(url_for('index'))


# ══════════════════════════════════════════════════════════════════════
#  Routes – Bank home (select mode)
# ══════════════════════════════════════════════════════════════════════

@app.route('/bank/home')
@require_bank
def bank_home():
    db = get_db()
    subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
    tags_rows = db.execute(
        "SELECT DISTINCT tag FROM questions WHERE tag != '' AND is_subquestion=0 ORDER BY tag"
    ).fetchall()
    tags = [r['tag'] for r in tags_rows]
    total = _count_answerable_questions(db)

    # Stats per type
    type_stats = {}
    for t in ['single_choice', 'multi_choice', 'true_false', 'fill_blank']:
        row = db.execute(
            "SELECT COUNT(*) FROM questions WHERE type=? AND type != 'composite'",
            (t,)
        ).fetchone()
        type_stats[t] = row[0]
    type_stats['composite_groups'] = db.execute(
        "SELECT COUNT(*) FROM questions WHERE type='composite' AND is_subquestion=0"
    ).fetchone()[0]

    bank_display = ''
    try:
        row = db.execute("SELECT value FROM bank_meta WHERE key='display_name'").fetchone()
        if row and row[0]:
            bank_display = row[0]
    except Exception:
        bank_display = session.get('current_bank', '')[:-3]
    return render_template('bank_home.html',
                           subjects=subjects, tags=tags, total=total,
                           type_stats=type_stats, bank_display=bank_display)


# ══════════════════════════════════════════════════════════════════════
#  Routes – Subject management
# ══════════════════════════════════════════════════════════════════════

@app.route('/api/subjects/add', methods=['POST'])
@require_bank
def api_add_subject():
    """AJAX endpoint to add a subject. Returns JSON."""
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'error': '请输入科目名称'}), 400
    db = get_db()
    try:
        cursor = db.execute('INSERT INTO subjects (name) VALUES (?)', (name,))
        db.commit()
        return jsonify({'id': cursor.lastrowid, 'name': name})
    except sqlite3.IntegrityError:
        return jsonify({'error': '科目已存在'}), 409


@app.route('/subjects/delete/<int:subject_id>', methods=['POST'])
@require_bank
def delete_subject(subject_id):
    db = get_db()
    db.execute('DELETE FROM subjects WHERE id=?', (subject_id,))
    db.commit()
    return redirect(url_for('bank_home'))


# ══════════════════════════════════════════════════════════════════════
#  Routes – Question CRUD
# ══════════════════════════════════════════════════════════════════════

@app.route('/questions')
@require_bank
def list_questions():
    db = get_db()
    sf = request.args.get('subject', '')
    tf = request.args.get('tag', '')
    typef = request.args.get('type', '')
    qf = request.args.get('q', '').strip()
    filters = {'subject': sf, 'tag': tf, 'type': typef, 'q': qf}
    question_rows = _query_top_level_questions(db, filters)[::-1]
    questions = []
    for row in question_rows:
        item = _question_row_to_dict(row)
        if item['type'] == 'composite':
            item['children'] = _fetch_children(db, item['id'])
            item['children_count'] = len(item['children'])
            item['total_score'] = _calculate_question_total_score(item)
        else:
            item['children'] = []
            item['children_count'] = 0
            item['effective_score'], item['uses_default_score'] = _resolve_question_score(item)
            item['total_score'] = item['effective_score']
        for child in item['children']:
            child['effective_score'], child['uses_default_score'] = _resolve_question_score(child)
        questions.append(item)
    subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
    tags_rows = db.execute(
        "SELECT DISTINCT tag FROM questions WHERE tag != '' AND is_subquestion=0 ORDER BY tag"
    ).fetchall()
    tags = [r['tag'] for r in tags_rows]
    current_bank = session.get('current_bank')
    other_banks = [bank for bank in get_bank_list() if bank['filename'] != current_bank]
    return render_template('questions.html',
                           questions=questions, subjects=subjects, tags=tags,
                           sf=sf, tf=tf, typef=typef, qf=qf, other_banks=other_banks)


@app.route('/questions/add', methods=['GET', 'POST'])
@require_bank
def add_question():
    db = get_db()
    if request.method == 'POST':
        try:
            qtype = request.form.get('type', 'single_choice')
            if qtype == 'composite':
                data = _parse_composite_form(request.form)
                cursor = db.execute(
                    """
                    INSERT INTO questions
                        (subject_id,tag,type,difficulty,content,options,answer,explanation,score,stem,is_subquestion)
                    VALUES (?,?,?,?,?,?,?,?,?,?,0)
                    """,
                    (data['subject_id'], data['tag'], 'composite', data['difficulty'],
                     data['content'], '[]', '', data['explanation'], None, data['content'])
                )
                parent_id = cursor.lastrowid
                for index, child in enumerate(data['children'], start=1):
                    db.execute(
                        """
                        INSERT INTO questions
                            (subject_id,tag,type,difficulty,content,options,answer,explanation,score,
                             parent_id,sort_order,stem,is_subquestion)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1)
                        """,
                        (data['subject_id'], data['tag'], child['type'], child['difficulty'],
                         child['content'], child['options'], child['answer'], child['explanation'], child['score'],
                         parent_id, index, data['content'])
                    )
            else:
                data = _parse_question_form(request.form)
                db.execute(
                    """
                    INSERT INTO questions
                        (subject_id,tag,type,difficulty,content,options,answer,explanation,score,stem,is_subquestion)
                    VALUES (?,?,?,?,?,?,?,?,?,?,0)
                    """,
                    (data['subject_id'], data['tag'], data['type'], data['difficulty'],
                     data['content'], data['options'], data['answer'], data['explanation'], data['score'], data['content'])
                )
        except ValueError as e:
            flash(str(e), 'error')
            subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
            return render_template('question_form.html', subjects=subjects, question=None)
        db.commit()
        flash('题目已添加', 'success')
        return redirect(url_for('list_questions'))

    subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
    return render_template('question_form.html', subjects=subjects, question=None)


@app.route('/questions/edit/<int:qid>', methods=['GET', 'POST'])
@require_bank
def edit_question(qid):
    db = get_db()
    current_question = _load_question_unit(db, qid)
    if not current_question:
        flash('题目不存在', 'error')
        return redirect(url_for('list_questions'))
    if request.method == 'POST':
        try:
            qtype = request.form.get('type', 'single_choice')
            if qtype == 'composite':
                data = _parse_composite_form(request.form)
                db.execute(
                    """
                    UPDATE questions
                    SET subject_id=?,tag=?,type='composite',difficulty=?,content=?,options='[]',
                        answer='',explanation=?,score=NULL,parent_id=NULL,sort_order=0,stem=?,is_subquestion=0
                    WHERE id=?
                    """,
                    (data['subject_id'], data['tag'], data['difficulty'], data['content'],
                     data['explanation'], data['content'], qid)
                )
                db.execute('DELETE FROM questions WHERE parent_id=?', (qid,))
                for index, child in enumerate(data['children'], start=1):
                    db.execute(
                        """
                        INSERT INTO questions
                            (subject_id,tag,type,difficulty,content,options,answer,explanation,score,
                             parent_id,sort_order,stem,is_subquestion)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1)
                        """,
                        (data['subject_id'], data['tag'], child['type'], child['difficulty'],
                         child['content'], child['options'], child['answer'], child['explanation'], child['score'],
                         qid, index, data['content'])
                    )
            else:
                data = _parse_question_form(request.form)
                db.execute(
                    """
                    UPDATE questions
                    SET subject_id=?,tag=?,type=?,difficulty=?,content=?,options=?,answer=?,explanation=?,score=?,
                        parent_id=NULL,sort_order=0,stem=?,is_subquestion=0
                    WHERE id=?
                    """,
                    (data['subject_id'], data['tag'], data['type'], data['difficulty'],
                     data['content'], data['options'], data['answer'], data['explanation'], data['score'],
                     data['content'], qid)
                )
                db.execute('DELETE FROM questions WHERE parent_id=?', (qid,))
        except ValueError as e:
            flash(str(e), 'error')
            subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
            return render_template('question_form.html', subjects=subjects, question=current_question)
        db.commit()
        flash('题目已更新', 'success')
        return redirect(url_for('list_questions'))

    subjects = db.execute('SELECT * FROM subjects ORDER BY name').fetchall()
    return render_template('question_form.html', subjects=subjects, question=current_question)


@app.route('/questions/delete/<int:qid>', methods=['POST'])
@require_bank
def delete_question(qid):
    db = get_db()
    db.execute('DELETE FROM questions WHERE parent_id=?', (qid,))
    db.execute('DELETE FROM questions WHERE id=?', (qid,))
    db.commit()
    flash('题目已删除', 'success')
    return redirect(url_for('list_questions'))


@app.route('/questions/view/<int:qid>')
@require_bank
def view_question(qid):
    db = get_db()
    q = _load_question_unit(db, qid)
    if not q:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(q)


def _parse_question_form(form):
    """Extract question fields from form data."""
    subject_id = form.get('subject_id') or None
    if subject_id:
        subject_id = int(subject_id)
    tag = form.get('tag', '').strip()
    qtype = form.get('type', 'single_choice')
    difficulty = int(form.get('difficulty', 1))
    raw_score = form.get('score', '')
    score = _coerce_positive_score(raw_score)
    if (raw_score or '').strip() and score is None:
        raise ValueError('题目分值必须是大于 0 的数字')
    content = form.get('content', '')
    explanation = form.get('explanation', '')

    option_map = {}
    for key in sorted(form):
        if key.startswith('option_'):
            option_map[key.replace('option_', '')] = form.get(key, '')
    options = _build_question_options(qtype, option_map)
    answer = _parse_question_answer(qtype, form.get('answer', ''))

    return {
        'subject_id': subject_id,
        'tag': tag,
        'type': qtype,
        'difficulty': difficulty,
        'score': score,
        'content': content,
        'options': json.dumps(options, ensure_ascii=False),
        'answer': answer,
        'explanation': explanation,
    }


def _parse_composite_form(form):
    subject_id = form.get('subject_id') or None
    if subject_id:
        subject_id = int(subject_id)

    tag = form.get('tag', '').strip()
    difficulty = int(form.get('difficulty', 1))
    content = form.get('content', '')
    explanation = form.get('explanation', '')
    payload = form.get('composite_payload', '[]')
    try:
        children_payload = json.loads(payload)
    except Exception as e:
        raise ValueError(f'组合题数据格式错误：{e}')

    children = []
    for index, raw in enumerate(children_payload, start=1):
        qtype = raw.get('type', 'single_choice')
        if qtype not in ANSWERABLE_TYPES:
            raise ValueError(f'第 {index} 个小题题型无效')
        child_content = (raw.get('content') or '').strip()
        if not child_content:
            raise ValueError(f'第 {index} 个小题题干不能为空')
        option_map = raw.get('options', {}) or {}
        options = _build_question_options(qtype, option_map)
        if qtype in ('single_choice', 'multi_choice') and len(options) < 2:
            raise ValueError(f'第 {index} 个小题至少需要两个选项')
        answer = _parse_question_answer(qtype, raw.get('answer', ''))
        raw_score = raw.get('score', '')
        score = _coerce_positive_score(raw_score)
        if str(raw_score or '').strip() and score is None:
            raise ValueError(f'第 {index} 个小题分值必须是大于 0 的数字')
        children.append({
            'type': qtype,
            'difficulty': int(raw.get('difficulty') or difficulty or 1),
            'score': score,
            'content': child_content,
            'options': json.dumps(options, ensure_ascii=False),
            'answer': answer,
            'explanation': raw.get('explanation', '').strip(),
        })

    if not content.strip():
        raise ValueError('组合题大题干不能为空')
    if not children:
        raise ValueError('组合题至少需要一个小题')

    return {
        'subject_id': subject_id,
        'tag': tag,
        'difficulty': difficulty,
        'content': content,
        'explanation': explanation,
        'children': children,
    }


# ══════════════════════════════════════════════════════════════════════
#  Routes – Excel import
# ══════════════════════════════════════════════════════════════════════

@app.route('/import', methods=['GET', 'POST'])
@require_bank
def import_questions():
    db = get_db()
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件', 'error')
            return redirect(url_for('import_questions'))
        f = request.files['file']
        if not f.filename.endswith(('.xlsx', '.xls')):
            flash('请上传 Excel 文件（.xlsx）', 'error')
            return redirect(url_for('import_questions'))

        # Auto-create subjects if needed
        existing = {r['name']: r['id'] for r in db.execute('SELECT * FROM subjects').fetchall()}

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            count = 0
            errors = []
            header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_map = {
                str(value).strip(): idx
                for idx, value in enumerate(header_values or [])
                if str(value or '').strip()
            }

            def get_cell(row, column_name, fallback_index=None):
                idx = header_map.get(column_name, fallback_index)
                if idx is None or idx >= len(row):
                    return ''
                return row[idx]

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    content_cell = get_cell(row, '题干', 4)
                    if not row or not str(content_cell or '').strip():
                        continue
                    subject_name = str(get_cell(row, '科目', 0) or '').strip()
                    tag = str(get_cell(row, '标签', 1) or '').strip()
                    qtype = str(get_cell(row, '题型', 2) or '').strip()
                    difficulty = int(get_cell(row, '难度', 3) or 1)
                    score = _coerce_positive_score(get_cell(row, '分值'))
                    content = str(content_cell or '').strip()
                    opt_a = str(get_cell(row, '选项A', 5) or '').strip()
                    opt_b = str(get_cell(row, '选项B', 6) or '').strip()
                    opt_c = str(get_cell(row, '选项C', 7) or '').strip()
                    opt_d = str(get_cell(row, '选项D', 8) or '').strip()
                    answer = str(get_cell(row, '正确答案', 9) or '').strip()
                    explanation = str(get_cell(row, '解析', 10) or '').strip()

                    # Validate type
                    type_map = {
                        'single_choice': 'single_choice',
                        'multi_choice': 'multi_choice',
                        'true_false': 'true_false',
                        'fill_blank': 'fill_blank',
                    }
                    qtype = type_map.get(qtype, 'single_choice')

                    # Ensure subject exists
                    subject_id = None
                    if subject_name:
                        if subject_name not in existing:
                            cursor = db.execute('INSERT INTO subjects (name) VALUES (?)', (subject_name,))
                            existing[subject_name] = cursor.lastrowid
                        subject_id = existing[subject_name]

                    # Build options
                    options = []
                    if qtype in ('single_choice', 'multi_choice'):
                        for label, val in [('A', opt_a), ('B', opt_b), ('C', opt_c), ('D', opt_d)]:
                            if val:
                                options.append({'label': label, 'content': val})
                    elif qtype == 'true_false':
                        options = [
                            {'label': 'A', 'content': '正确'},
                            {'label': 'B', 'content': '错误'},
                        ]

                    db.execute(
                        "INSERT INTO questions (subject_id,tag,type,difficulty,content,options,answer,explanation,score) "
                        "VALUES (?,?,?,?,?,?,?,?,?)",
                        (subject_id, tag, qtype, difficulty, content,
                         json.dumps(options, ensure_ascii=False), answer, explanation, score)
                    )
                    count += 1
                except Exception as e:
                    errors.append(f'第{i}行：{e}')

            db.commit()
            msg = f'成功导入 {count} 道题目'
            if errors:
                msg += f'，{len(errors)} 行出错'
            flash(msg, 'success' if not errors else 'warning')
            if errors:
                for e in errors[:5]:
                    flash(e, 'error')
        except Exception as e:
            flash(f'导入失败：{e}', 'error')

        return redirect(url_for('list_questions'))

    return render_template('import.html')


# ══════════════════════════════════════════════════════════════════════
#  Routes – Export & JSON import
# ══════════════════════════════════════════════════════════════════════

@app.route('/export')
@require_bank
def export_bank():
    """Export current bank as a ZIP package with JSON + referenced assets."""
    db = get_db()
    bank_name = _get_bank_display_name(db, 'bank')
    filters = _build_export_filters(request.args)

    subjects = [{'id': r['id'], 'name': r['name']}
                for r in db.execute('SELECT * FROM subjects ORDER BY id').fetchall()]

    subject_name = ''
    if filters['subject']:
        row = db.execute('SELECT name FROM subjects WHERE id=?', (filters['subject'],)).fetchone()
        if row and row[0]:
            subject_name = row[0]

    selected_rows = _query_top_level_questions(db, filters)
    questions = []
    for row in selected_rows:
        questions.append(_serialize_question_export(_load_question_unit(db, row['id'])))

    data = {
        'format': 'family-exam-helper',
        'version': 2,
        'bank_name': bank_name,
        'subjects': subjects,
        'questions': questions,
        'filters': {
            'subject': subject_name,
            'tag': filters['tag'],
            'type': filters['type'],
        },
        'exported_at': datetime.now().isoformat(),
    }
    file_name = _build_export_file_name(
        bank_name,
        {
            'subject_name': subject_name,
            'tag': filters['tag'],
            'type_label': QUESTION_TYPE_LABELS.get(filters['type'], filters['type']),
        }
    )
    return _build_export_package(data, file_name)


@app.route('/import/json', methods=['GET', 'POST'])
@require_bank
def import_json():
    """Import questions from a previously exported JSON or ZIP package."""
    if request.method == 'GET':
        return render_template('import_json.html')

    if 'file' not in request.files:
        flash('请选择文件', 'error')
        return redirect(url_for('import_json'))

    f = request.files['file']
    if not f.filename.lower().endswith(('.json', '.zip')):
        flash('请上传导出包（.zip）或 JSON 文件（.json）', 'error')
        return redirect(url_for('import_json'))

    try:
        data, imported_assets = _load_import_payload(f)
    except Exception as e:
        flash(f'文件读取失败：{e}', 'error')
        return redirect(url_for('import_json'))

    if data.get('format') != 'family-exam-helper':
        flash('不是有效的题库导出文件', 'error')
        return redirect(url_for('import_json'))

    db = get_db()

    # Build existing subject map (name -> id)
    existing = {r['name']: r['id'] for r in db.execute('SELECT * FROM subjects').fetchall()}

    # Import subjects (merge – skip existing)
    for s in data.get('subjects', []):
        name = s.get('name', '').strip()
        if name and name not in existing:
            cursor = db.execute('INSERT INTO subjects (name) VALUES (?)', (name,))
            existing[name] = cursor.lastrowid

    # Import questions
    count = 0
    errors = []
    for i, q in enumerate(data.get('questions', []), start=1):
        try:
            if imported_assets:
                q = _inline_imported_asset_refs(q, imported_assets, _current_bank_filename())
            _import_question_payload(db, existing, q)
            count += 1
        except Exception as e:
            errors.append(f'第{i}题：{e}')

    db.commit()

    msg = f'成功导入 {count} 道题目'
    if errors:
        msg += f'，{len(errors)} 题出错'
    flash(msg, 'success' if not errors else 'warning')
    if errors:
        for e in errors[:5]:
            flash(e, 'error')
    return redirect(url_for('list_questions'))


@app.route('/import/template', methods=['POST'])
@require_bank
def import_template():
    """Import questions from the public standard template package."""
    if 'file' not in request.files:
        flash('请选择文件', 'error')
        return redirect(url_for('import_questions'))

    f = request.files['file']
    if not f.filename.lower().endswith(('.json', '.zip')):
        flash('请上传标准模板文件（.zip）或模板 JSON（.json）', 'error')
        return redirect(url_for('import_questions'))

    try:
        data, imported_assets = _load_standard_template_payload(f)
    except Exception as e:
        flash(f'模板读取失败：{e}', 'error')
        return redirect(url_for('import_questions'))

    if data.get('format') != 'family-exam-helper-template':
        flash('不是有效的标准导入模板文件', 'error')
        return redirect(url_for('import_questions'))

    db = get_db()
    existing = {r['name']: r['id'] for r in db.execute('SELECT * FROM subjects').fetchall()}

    for name in _extract_template_subject_names(data):
        if name and name not in existing:
            cursor = db.execute('INSERT INTO subjects (name) VALUES (?)', (name,))
            existing[name] = cursor.lastrowid

    count = 0
    errors = []
    for i, q in enumerate(data.get('questions', []), start=1):
        try:
            if imported_assets:
                q = _inline_template_imported_asset_refs(q, imported_assets, _current_bank_filename())
            _import_question_payload(db, existing, q)
            count += 1
        except Exception as e:
            errors.append(f'第{i}题：{e}')

    db.commit()

    msg = f'成功导入 {count} 道题目'
    if errors:
        msg += f'，{len(errors)} 题出错'
    flash(msg, 'success' if not errors else 'warning')
    if errors:
        for e in errors[:5]:
            flash(e, 'error')
    return redirect(url_for('list_questions'))


@app.route('/questions/export/<int:qid>/to-bank', methods=['POST'])
@require_bank
def export_question_to_bank(qid):
    target_bank = request.form.get('target_bank', '').strip()
    current_bank = session.get('current_bank')
    if not target_bank:
        flash('请选择目标题库', 'error')
        return redirect(request.referrer or url_for('list_questions'))
    if target_bank == current_bank:
        flash('不能导出到当前题库', 'error')
        return redirect(request.referrer or url_for('list_questions'))

    target_path = _bank_path(target_bank)
    if not target_path or not os.path.exists(target_path):
        flash('目标题库不存在', 'error')
        return redirect(request.referrer or url_for('list_questions'))

    source_db = get_db()
    question = _load_question_unit(source_db, qid)
    if not question:
        flash('题目不存在', 'error')
        return redirect(request.referrer or url_for('list_questions'))

    target_db = sqlite3.connect(target_path, detect_types=sqlite3.PARSE_DECLTYPES)
    target_db.row_factory = sqlite3.Row
    try:
        migrate_bank_schema(target_db)
        existing = {r['name']: r['id'] for r in target_db.execute('SELECT * FROM subjects').fetchall()}
        exported_question = _serialize_question_export(question)
        mapped_question = _remap_question_assets(exported_question, target_bank)
        _import_question_payload(target_db, existing, mapped_question)
        target_db.commit()
        target_name = _get_bank_display_name(target_db, target_bank[:-3])
    finally:
        target_db.close()

    flash(f'题目已导出到「{target_name}」', 'success')
    return redirect(request.referrer or url_for('list_questions'))


# ══════════════════════════════════════════════════════════════════════
#  Routes – Practice mode
# ══════════════════════════════════════════════════════════════════════

@app.route('/practice/start', methods=['POST'])
@require_bank
def practice_start():
    db = get_db()
    qids = _select_question_ids(db, request.form)
    if not qids:
        flash('没有符合条件的题目', 'error')
        return redirect(url_for('bank_home'))
    session['practice_qids'] = qids
    session['practice_index'] = 0
    session['practice_answers'] = {}
    session['practice_start'] = time.time()
    return redirect(url_for('practice_question'))


@app.route('/practice')
@require_bank
def practice_question():
    qids = session.get('practice_qids', [])
    idx = session.get('practice_index', 0)
    if not qids or idx >= len(qids):
        return redirect(url_for('practice_result'))

    db = get_db()
    q = _load_question_unit(db, qids[idx])
    if not q:
        return redirect(url_for('practice_result'))

    total = len(qids)
    return render_template('practice.html',
                           question=q, index=idx, total=total,
                           options=q['options'],
                           saved_answer=session.get('practice_answers', {}).get(str(q['id']), ''),
                           saved_answers=session.get('practice_answers', {}).get(str(q['id']), {}))


@app.route('/practice/next', methods=['POST'])
@require_bank
def practice_next():
    qid = request.form.get('qid')
    unit_type = request.form.get('unit_type', '')
    answer = request.form.get('answer', '').strip()
    action = request.form.get('action', 'next')

    if qid:
        answers = session.get('practice_answers', {})
        db = get_db()
        question = _load_question_unit(db, int(qid))
        if question and (unit_type == 'composite' or question['type'] == 'composite'):
            answers[str(qid)] = _serialize_answer_for_session(_extract_composite_answers(request.form, question['children']))
        else:
            answers[str(qid)] = answer if question is None else _parse_question_answer(question['type'], answer)
        session['practice_answers'] = answers

    if action == 'prev':
        session['practice_index'] = max(0, session.get('practice_index', 0) - 1)
    elif action == 'finish':
        return redirect(url_for('practice_result'))
    else:
        idx = session.get('practice_index', 0) + 1
        session['practice_index'] = idx

    return redirect(url_for('practice_question'))


@app.route('/practice/result')
@require_bank
def practice_result():
    qids = session.get('practice_qids', [])
    answers = session.get('practice_answers', {})
    if not qids:
        return redirect(url_for('bank_home'))

    db = get_db()
    questions = []
    correct = 0
    total_items = 0
    for qid in qids:
        q = _load_question_unit(db, qid)
        if q:
            user_ans = answers.get(str(qid), '')
            result = _build_result_entry(q, user_ans)
            correct += result['group_correct']
            total_items += result['group_total']
            questions.append(result)

    duration = int(time.time() - session.get('practice_start', time.time()))
    # Clear session
    for key in ['practice_qids', 'practice_index', 'practice_answers', 'practice_start']:
        session.pop(key, None)

    return render_template('result.html',
                           mode='practice', questions=questions,
                           correct=correct, total=total_items,
                           duration=duration)


# ══════════════════════════════════════════════════════════════════════
#  Routes – Exam mode
# ══════════════════════════════════════════════════════════════════════

@app.route('/exam/start', methods=['POST'])
@require_bank
def exam_start():
    db = get_db()
    qids = _select_question_ids(db, request.form)
    if not qids:
        flash('没有符合条件的题目', 'error')
        return redirect(url_for('bank_home'))
    session['exam_qids'] = qids
    session['exam_index'] = 0
    session['exam_answers'] = {}
    session['exam_start'] = time.time()
    return redirect(url_for('exam_question'))


@app.route('/exam')
@require_bank
def exam_question():
    qids = session.get('exam_qids', [])
    idx = session.get('exam_index', 0)
    if not qids or idx >= len(qids):
        return redirect(url_for('exam_submit'))

    db = get_db()
    q = _load_question_unit(db, qids[idx])
    if not q:
        return redirect(url_for('exam_submit'))

    total = len(qids)
    return render_template('exam.html',
                           question=q, index=idx, total=total,
                           options=q['options'],
                           saved_answer=session.get('exam_answers', {}).get(str(q['id']), ''),
                           saved_answers=session.get('exam_answers', {}).get(str(q['id']), {}))


@app.route('/exam/next', methods=['POST'])
@require_bank
def exam_next():
    qid = request.form.get('qid')
    unit_type = request.form.get('unit_type', '')
    answer = request.form.get('answer', '').strip()
    action = request.form.get('action', 'next')

    if qid:
        answers = session.get('exam_answers', {})
        db = get_db()
        question = _load_question_unit(db, int(qid))
        if question and (unit_type == 'composite' or question['type'] == 'composite'):
            answers[str(qid)] = _serialize_answer_for_session(_extract_composite_answers(request.form, question['children']))
        else:
            answers[str(qid)] = answer if question is None else _parse_question_answer(question['type'], answer)
        session['exam_answers'] = answers

    if action == 'prev':
        session['exam_index'] = max(0, session.get('exam_index', 0) - 1)
    elif action == 'finish':
        return redirect(url_for('exam_submit'))
    else:
        idx = session.get('exam_index', 0) + 1
        session['exam_index'] = idx

    return redirect(url_for('exam_question'))


@app.route('/exam/submit')
@require_bank
def exam_submit():
    qids = session.get('exam_qids', [])
    answers = session.get('exam_answers', {})
    if not qids:
        return redirect(url_for('bank_home'))

    db = get_db()
    questions = []
    correct = 0
    total_items = 0
    raw_score = 0.0
    total_score = 0.0

    for qid in qids:
        q = _load_question_unit(db, qid)
        if q:
            user_ans = answers.get(str(qid), '')
            result = _build_result_entry(q, user_ans)
            correct += result['group_correct']
            total_items += result['group_total']
            raw_score += result['score']
            total_score += result.get('max_score', 0)
            questions.append(result)

    normalized_score = round(raw_score / total_score * 100) if total_score else 0
    duration = int(time.time() - session.get('exam_start', time.time()))

    # Save exam record
    try:
        db.execute(
            "INSERT INTO exam_records (mode,total_questions,correct_count,score,total_score,duration_seconds,answers) "
            "VALUES (?,?,?,?,?,?,?)",
            ('exam', total_items, correct, raw_score, total_score,
             duration, json.dumps(answers, ensure_ascii=False))
        )
        db.commit()
    except Exception:
        pass

    # Clear session
    for key in ['exam_qids', 'exam_index', 'exam_answers', 'exam_start']:
        session.pop(key, None)

    return render_template('result.html',
                           mode='exam', questions=questions,
                           correct=correct, total=total_items,
                           score=raw_score, total_score=total_score,
                           normalized_score=normalized_score,
                           duration=duration)


# ══════════════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════════════

def _select_question_ids(db, form):
    """Build question ID list from form filters."""
    subject_id = form.get('subject_id', '')
    tag = form.get('tag', '')
    count = int(form.get('count', 0) or 0)

    q = "SELECT id FROM questions WHERE is_subquestion=0"
    params = []
    if subject_id:
        q += " AND subject_id=?"
        params.append(int(subject_id))
    if tag:
        q += " AND tag=?"
        params.append(tag)
    q += " ORDER BY RANDOM()"
    if count > 0:
        q += " LIMIT ?"
        params.append(count)

    rows = db.execute(q, params).fetchall()
    return [r['id'] for r in rows]


def _check_answer(qtype, correct_answer, user_answer):
    """Check if user answer matches correct answer."""
    if not user_answer:
        return False
    user_answer = user_answer.strip()
    correct_answer = correct_answer.strip()

    if qtype == 'fill_blank':
        # Multiple acceptable answers separated by |
        acceptable = [a.strip().lower() for a in correct_answer.split('|')]
        return user_answer.lower() in acceptable
    elif qtype == 'multi_choice':
        # All correct options must be selected, no extras
        correct_set = {part for part in _normalize_multi_answer(correct_answer).split(',') if part}
        user_set = {part for part in _normalize_multi_answer(user_answer).split(',') if part}
        return correct_set == user_set
    else:
        # single_choice, true_false
        return user_answer.upper() == correct_answer.upper()


# ══════════════════════════════════════════════════════════════════════
#  Run
# ══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
